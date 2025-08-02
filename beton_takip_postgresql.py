import psycopg2
import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox
from psycopg2.extras import RealDictCursor
import configparser
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from decimal import Decimal

# === EXCEL KAYIT FONKSİYONLARI ===
def excel_kayit_olustur(islem_tipi, veri_dict):
    """Her işlem için otomatik Excel kaydı oluşturur"""
    try:
        # Kayıt klasörünü oluştur
        kayit_klasoru = "excel_kayitlari"
        os.makedirs(kayit_klasoru, exist_ok=True)
        
        # Dosya adını belirle
        dosya_adi = f"{kayit_klasoru}/{islem_tipi}.xlsx"
        
        # Stil tanımlamaları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Dosya var mı kontrol et
        if os.path.exists(dosya_adi):
            # Mevcut dosyayı aç
            wb = openpyxl.load_workbook(dosya_adi)
            ws = wb.active
        else:
            # Yeni dosya oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = islem_tipi
            
            # Header'ları ekle
            headers = list(veri_dict.keys())
            ws.append(headers)
            
            # Header'ları formatla
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        # Yeni veriyi ekle
        ws.append(list(veri_dict.values()))
        
        # Son satırı formatla
        son_satir = ws.max_row
        for cell in ws[son_satir]:
            cell.border = border
        
        # Sütun genişliklerini ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Dosyayı kaydet
        wb.save(dosya_adi)
        print(f"Excel kaydı oluşturuldu: {dosya_adi}")
        
    except Exception as e:
        print(f"Excel kayıt hatası: {str(e)}")

# === VERİTABANI BAĞLANTI AYARLARI ===
class DatabaseConfig:
    def __init__(self):
        self.config_file = "db_config.ini"
        self.load_config()
    
    def load_config(self):
        """Konfigürasyon dosyasından veritabanı ayarlarını yükle"""
        config = configparser.ConfigParser()
        
        if not os.path.exists(self.config_file):
            self.create_default_config()
        
        config.read(self.config_file)
        
        self.host = config.get('database', 'host', fallback='localhost')
        self.port = config.get('database', 'port', fallback='5432')
        self.database = config.get('database', 'database', fallback='beton_takip')
        self.username = config.get('database', 'username', fallback='postgres')
        self.password = config.get('database', 'password', fallback='password')
    
    def create_default_config(self):
        """Varsayılan konfigürasyon dosyası oluştur"""
        config = configparser.ConfigParser()
        config['database'] = {
            'host': 'localhost',
            'port': '5432',
            'database': 'beton_takip',
            'username': 'postgres',
            'password': 'password'
        }
        
        with open(self.config_file, 'w') as configfile:
            config.write(configfile)
        
        messagebox.showinfo("Konfigürasyon", 
            f"{self.config_file} dosyası oluşturuldu. Veritabanı bağlantı ayarlarınızı düzenleyin.")

# === VERİTABANI YÖNETİCİSİ ===
class DatabaseManager:
    def __init__(self):
        self.config = DatabaseConfig()
        self.connection = None
        self.connect()
        self.create_tables()
    
    def connect(self):
        """Veritabanına bağlan"""
        try:
            self.connection = psycopg2.connect(
                host=self.config.host,
                port=self.config.port,
                database=self.config.database,
                user=self.config.username,
                password=self.config.password
            )
            self.connection.autocommit = True
            print("Veritabanı bağlantısı başarılı!")
        except Exception as e:
            messagebox.showerror("Veritabanı Hatası", 
                f"Veritabanına bağlanılamadı: {str(e)}\n\ndb_config.ini dosyasını kontrol edin.")
            raise
    
    def create_tables(self):
        """Gerekli tabloları oluştur"""
        cursor = self.connection.cursor()
        
        # Stok tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS stok (
                id SERIAL PRIMARY KEY,
                malzeme VARCHAR(255) UNIQUE NOT NULL,
                miktar_kg DECIMAL(10,2) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Alışlar tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS alislar (
                id SERIAL PRIMARY KEY,
                malzeme VARCHAR(255) NOT NULL,
                miktar_kg DECIMAL(10,2) NOT NULL,
                birim_fiyat DECIMAL(10,2) NOT NULL,
                toplam_tutar DECIMAL(10,2) NOT NULL,
                tarih DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Ürünler tablosu (reçeteler)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS urunler (
                id SERIAL PRIMARY KEY,
                urun VARCHAR(255) NOT NULL,
                malzeme VARCHAR(255) NOT NULL,
                yuzde DECIMAL(5,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Üretimler tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS uretimler (
                id SERIAL PRIMARY KEY,
                urun VARCHAR(255) NOT NULL,
                gramaj_kg DECIMAL(10,2) NOT NULL,
                tarih DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Satışlar tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS satislar (
                id SERIAL PRIMARY KEY,
                urun VARCHAR(255) NOT NULL,
                musteri VARCHAR(255) NOT NULL,
                miktar_kg DECIMAL(10,2) NOT NULL,
                satis_fiyat DECIMAL(10,2) NOT NULL,
                toplam_satis DECIMAL(10,2) NOT NULL,
                net_kar DECIMAL(10,2),
                tarih DATE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # İade/Hurda tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS iadeler (
                id SERIAL PRIMARY KEY,
                tarih DATE NOT NULL,
                tip VARCHAR(50) NOT NULL,
                urun VARCHAR(255) NOT NULL,
                miktar DECIMAL(10,2) NOT NULL,
                sebep TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Taş gelir-gider tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tas_gelir_gider (
                id SERIAL PRIMARY KEY,
                tarih DATE NOT NULL,
                tip VARCHAR(50) NOT NULL,
                aciklama VARCHAR(255) NOT NULL,
                birim VARCHAR(50),
                birim_fiyat DECIMAL(10,2) NOT NULL,
                miktar DECIMAL(10,2) NOT NULL,
                toplam_tutar DECIMAL(10,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Beton gelir-gider tablosu
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS beton_gelir_gider (
                id SERIAL PRIMARY KEY,
                tarih DATE NOT NULL,
                tip VARCHAR(50) NOT NULL,
                aciklama VARCHAR(255) NOT NULL,
                birim VARCHAR(50),
                birim_fiyat DECIMAL(10,2) NOT NULL,
                miktar DECIMAL(10,2) NOT NULL,
                toplam_tutar DECIMAL(10,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.close()
    
    def execute_query(self, query, params=None):
        """SQL sorgusu çalıştır"""
        cursor = self.connection.cursor(cursor_factory=RealDictCursor)
        cursor.execute(query, params)
        return cursor
    
    def fetch_all(self, query, params=None):
        """Tüm sonuçları getir"""
        cursor = self.execute_query(query, params)
        result = cursor.fetchall()
        cursor.close()
        return result
    
    def fetch_one(self, query, params=None):
        """Tek sonuç getir"""
        cursor = self.execute_query(query, params)
        result = cursor.fetchone()
        cursor.close()
        return result
    
    def insert(self, table, data):
        """Veri ekle"""
        columns = ', '.join(data.keys())
        placeholders = ', '.join(['%s'] * len(data))
        query = f"INSERT INTO {table} ({columns}) VALUES ({placeholders})"
        
        cursor = self.connection.cursor()
        cursor.execute(query, list(data.values()))
        cursor.close()
    
    def update(self, table, data, where_clause, where_params):
        """Veri güncelle"""
        set_clause = ', '.join([f"{k} = %s" for k in data.keys()])
        query = f"UPDATE {table} SET {set_clause} WHERE {where_clause}"
        
        cursor = self.connection.cursor()
        cursor.execute(query, list(data.values()) + where_params)
        cursor.close()

# Global veritabanı yöneticisi
db = None

try:
    db = DatabaseManager()
except:
    exit()

# === YARDIMCI FONKSİYONLAR ===
def get_malzemeler():
    """Stokta bulunan malzemeleri getir"""
    result = db.fetch_all("SELECT DISTINCT malzeme FROM stok ORDER BY malzeme")
    return [row['malzeme'] for row in result]

def get_urunler():
    """Tanımlı ürünleri getir"""
    result = db.fetch_all("SELECT DISTINCT urun FROM urunler ORDER BY urun")
    return [row['urun'] for row in result]

def get_malzeme_ve_urunler():
    """Malzeme ve ürünlerin birleşik listesi"""
    malzemeler = get_malzemeler()
    urunler = get_urunler()
    return sorted(set(malzemeler + urunler))

# === ARAYÜZ BAŞLAT ===
root = tk.Tk()
root.title("Beton Parke Takip Sistemi - PostgreSQL")
root.geometry("800x600")
notebook = ttk.Notebook(root)
notebook.pack(expand=True, fill="both")

# === STOK GİRİŞİ SEKMESİ ===
def stok_girisi():
    try:
        malzeme = entry_malzeme.get()
        miktar = Decimal(str(entry_miktar.get()))
        fiyat = Decimal(str(entry_fiyat.get()))
        tarih = datetime.now().date()
        toplam_tutar = miktar * fiyat

        # Alış kaydı ekle
        alis_data = {
            'malzeme': malzeme,
            'miktar_kg': miktar,
            'birim_fiyat': fiyat,
            'toplam_tutar': toplam_tutar,
            'tarih': tarih
        }
        db.insert('alislar', alis_data)

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Malzeme': malzeme,
            'Miktar (kg)': float(miktar),
            'Birim Fiyat': float(fiyat),
            'Toplam Tutar': float(toplam_tutar),
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Alislar", excel_data)

        # Stok güncelle
        existing = db.fetch_one("SELECT * FROM stok WHERE malzeme = %s", [malzeme])
        if existing:
            new_miktar = existing['miktar_kg'] + miktar
            db.update('stok', {'miktar_kg': new_miktar, 'updated_at': datetime.now()}, 
                     'malzeme = %s', [malzeme])
        else:
            db.insert('stok', {'malzeme': malzeme, 'miktar_kg': miktar})

        messagebox.showinfo("Başarılı", "Stok girişi kaydedildi ve Excel'e aktarıldı.")
        entry_malzeme.delete(0, tk.END)
        entry_miktar.delete(0, tk.END)
        entry_fiyat.delete(0, tk.END)
        
        # Combobox'ları güncelle
        guncelle_comboboxlar()
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f1 = ttk.Frame(notebook)
notebook.add(f1, text="Stok Girişi")
tk.Label(f1, text="Malzeme: ").grid(row=0, column=0, padx=5, pady=5)
entry_malzeme = tk.Entry(f1)
entry_malzeme.grid(row=0, column=1, padx=5, pady=5)
tk.Label(f1, text="Miktar (kg): ").grid(row=1, column=0, padx=5, pady=5)
entry_miktar = tk.Entry(f1)
entry_miktar.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f1, text="Birim Fiyat: ").grid(row=2, column=0, padx=5, pady=5)
entry_fiyat = tk.Entry(f1)
entry_fiyat.grid(row=2, column=1, padx=5, pady=5)
tk.Button(f1, text="Kaydet", command=stok_girisi).grid(row=3, columnspan=2, pady=10)
tk.Label(f1, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=4, columnspan=2, pady=5)

# === ÜRÜN REÇETESİ TANIMI SEKMESİ ===
recete_gecici = []

def receteye_malzeme_ekle():
    malzeme = combo_urun_malzeme.get()
    try:
        yuzde = Decimal(str(entry_urun_yuzde.get()))
        urun_adi = entry_urun.get()
        if not urun_adi or not malzeme:
            raise ValueError("Ürün adı ve malzeme seçilmelidir.")
        
        recete_gecici.append((urun_adi, malzeme, yuzde))
        liste_kutu.insert(tk.END, f"{malzeme} - %{float(yuzde)}")
        combo_urun_malzeme.set("")
        entry_urun_yuzde.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))

def recete_kaydet():
    if not recete_gecici:
        messagebox.showwarning("Uyarı", "Hiç malzeme eklenmedi.")
        return
    
    try:
        for urun, malzeme, yuzde in recete_gecici:
            db.insert('urunler', {
                'urun': urun,
                'malzeme': malzeme,
                'yuzde': yuzde
            })

            # Excel kaydı oluştur
            excel_data = {
                'Tarih': datetime.now().strftime("%Y-%m-%d"),
                'Urun': urun,
                'Malzeme': malzeme,
                'Yuzde': float(yuzde),
                'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            excel_kayit_olustur("Urun_Receteleri", excel_data)
        
        messagebox.showinfo("Başarılı", "Ürün reçetesi kaydedildi ve Excel'e aktarıldı.")
        entry_urun.delete(0, tk.END)
        liste_kutu.delete(0, tk.END)
        recete_gecici.clear()
        guncelle_comboboxlar()
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f2 = ttk.Frame(notebook)
notebook.add(f2, text="Ürün Tanımı")
tk.Label(f2, text="Ürün Adı: ").grid(row=0, column=0, padx=5, pady=5)
entry_urun = tk.Entry(f2, width=30)
entry_urun.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
tk.Label(f2, text="Malzeme: ").grid(row=1, column=0, padx=5, pady=5)
combo_urun_malzeme = ttk.Combobox(f2, values=get_malzemeler(), state="readonly")
combo_urun_malzeme.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f2, text="Yüzde: ").grid(row=1, column=2, padx=5, pady=5)
entry_urun_yuzde = tk.Entry(f2)
entry_urun_yuzde.grid(row=1, column=3, padx=5, pady=5)
tk.Button(f2, text="Malzeme Ekle", command=receteye_malzeme_ekle).grid(row=2, column=0, columnspan=4, pady=5)
liste_kutu = tk.Listbox(f2, width=60)
liste_kutu.grid(row=3, column=0, columnspan=4, padx=5, pady=5)
tk.Button(f2, text="Reçeteyi Kaydet", command=recete_kaydet).grid(row=4, column=0, columnspan=4, pady=5)
tk.Label(f2, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=5, column=0, columnspan=4, pady=5)

# === ÜRETİM SEKMESİ ===
def uretim_yap():
    try:
        urun = combo_uretim_urun.get()
        gramaj = Decimal(str(entry_uretim_gramaj.get()))
        tarih = datetime.now().date()

        # Ürün reçetesini kontrol et
        recete = db.fetch_all("SELECT * FROM urunler WHERE urun = %s", [urun])
        if not recete:
            raise ValueError("Bu ürün için reçete tanımı yok.")

        # Kullanılan malzemeler listesi
        kullanilan_malzemeler = []

        # Stok kontrolü ve düşürme
        for row in recete:
            malzeme = row['malzeme']
            oran = Decimal(str(row['yuzde'])) / Decimal('100')
            gereken = gramaj * oran
            
            stok_row = db.fetch_one("SELECT * FROM stok WHERE malzeme = %s", [malzeme])
            if not stok_row:
                raise ValueError(f"{malzeme} stokta yok.")
            
            mevcut = Decimal(str(stok_row['miktar_kg']))
            if mevcut < gereken:
                raise ValueError(f"{malzeme} için yeterli stok yok. Mevcut: {mevcut}, Gereken: {gereken}")
            
            # Stoktan düş
            db.update('stok', 
                     {'miktar_kg': mevcut - gereken, 'updated_at': datetime.now()},
                     'malzeme = %s', [malzeme])
            
            kullanilan_malzemeler.append(f"{malzeme}: {float(gereken):.2f} kg")

        # Üretim kaydı ekle
        db.insert('uretimler', {
            'urun': urun,
            'gramaj_kg': gramaj,
            'tarih': tarih
        })

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Urun': urun,
            'Gramaj (kg)': float(gramaj),
            'Kullanilan Malzemeler': " | ".join(kullanilan_malzemeler),
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Uretimler", excel_data)

        messagebox.showinfo("Başarılı", "Üretim kaydedildi ve Excel'e aktarıldı.")
        combo_uretim_urun.set("")
        entry_uretim_gramaj.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f3 = ttk.Frame(notebook)
notebook.add(f3, text="Üretim")
tk.Label(f3, text="Ürün: ").grid(row=0, column=0, padx=5, pady=5)
combo_uretim_urun = ttk.Combobox(f3, values=get_urunler(), state="readonly")
combo_uretim_urun.grid(row=0, column=1, padx=5, pady=5)
tk.Label(f3, text="Gramaj (kg): ").grid(row=1, column=0, padx=5, pady=5)
entry_uretim_gramaj = tk.Entry(f3)
entry_uretim_gramaj.grid(row=1, column=1, padx=5, pady=5)
tk.Button(f3, text="Üretimi Kaydet", command=uretim_yap).grid(row=2, column=0, columnspan=2, pady=10)
tk.Label(f3, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=3, columnspan=2, pady=5)

# === SATIŞ SEKMESİ ===
def satis_kaydet():
    try:
        urun = combo_satis_urun.get()
        musteri = entry_satis_musteri.get()
        miktar = Decimal(str(entry_satis_miktar.get()))
        fiyat = Decimal(str(entry_satis_fiyat.get()))
        tarih = datetime.now().date()
        toplam_satis = miktar * fiyat

        kdv_orani = Decimal('0.20')
        
        # Maliyet hesapla
        recete = db.fetch_all("SELECT * FROM urunler WHERE urun = %s", [urun])
        toplam_maliyet = Decimal('0')
        maliyet_detay = []
        
        for row in recete:
            malzeme = row['malzeme']
            oran = Decimal(str(row['yuzde'])) / Decimal('100')
            gereken_miktar = miktar * oran
            
            # Son alış fiyatını al
            alis_row = db.fetch_one(
                "SELECT birim_fiyat FROM alislar WHERE malzeme = %s ORDER BY tarih DESC LIMIT 1",
                [malzeme]
            )
            if alis_row:
                birim_fiyat = Decimal(str(alis_row['birim_fiyat']))
                malzeme_maliyet = gereken_miktar * birim_fiyat
                toplam_maliyet += malzeme_maliyet
                maliyet_detay.append(f"{malzeme}: {float(malzeme_maliyet):.2f} TL")

        net_kar = (fiyat * miktar / (Decimal('1') + kdv_orani)) - toplam_maliyet

        # Satış kaydı ekle
        satis_data = {
            'urun': urun,
            'musteri': musteri,
            'miktar_kg': miktar,
            'satis_fiyat': fiyat,
            'toplam_satis': toplam_satis,
            'net_kar': net_kar,
            'tarih': tarih
        }
        db.insert('satislar', satis_data)

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Urun': urun,
            'Musteri': musteri,
            'Miktar (kg)': float(miktar),
            'Birim Fiyat': float(fiyat),
            'Toplam Satis': float(toplam_satis),
            'Toplam Maliyet': float(toplam_maliyet),
            'Net Kar': float(net_kar),
            'Maliyet Detay': " | ".join(maliyet_detay),
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Satislar", excel_data)

        messagebox.showinfo("Başarılı", "Satış kaydedildi ve Excel'e aktarıldı.")
        combo_satis_urun.set("")
        entry_satis_musteri.delete(0, tk.END)
        entry_satis_miktar.delete(0, tk.END)
        entry_satis_fiyat.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f4 = ttk.Frame(notebook)
notebook.add(f4, text="Satış")
tk.Label(f4, text="Ürün: ").grid(row=0, column=0, padx=5, pady=5)
combo_satis_urun = ttk.Combobox(f4, values=get_urunler(), state="readonly")
combo_satis_urun.grid(row=0, column=1, padx=5, pady=5)
tk.Label(f4, text="Müşteri: ").grid(row=1, column=0, padx=5, pady=5)
entry_satis_musteri = tk.Entry(f4)
entry_satis_musteri.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f4, text="Miktar (kg): ").grid(row=2, column=0, padx=5, pady=5)
entry_satis_miktar = tk.Entry(f4)
entry_satis_miktar.grid(row=2, column=1, padx=5, pady=5)
tk.Label(f4, text="Satış Fiyatı: ").grid(row=3, column=0, padx=5, pady=5)
entry_satis_fiyat = tk.Entry(f4)
entry_satis_fiyat.grid(row=3, column=1, padx=5, pady=5)
tk.Button(f4, text="Satışı Kaydet", command=satis_kaydet).grid(row=4, column=0, columnspan=2, pady=10)
tk.Label(f4, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=5, columnspan=2, pady=5)

# === İADE/HURDA SEKMESİ ===
def iade_kaydet():
    try:
        urun = combo_iade_urun.get()
        miktar = Decimal(str(entry_iade_miktar.get()))
        sebep = entry_iade_sebep.get()
        tip = combo_iade_tip.get()
        tarih = datetime.now().date()

        # İade kaydı ekle
        db.insert('iadeler', {
            'tarih': tarih,
            'tip': tip,
            'urun': urun,
            'miktar': miktar,
            'sebep': sebep
        })

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Tip': tip,
            'Urun/Malzeme': urun,
            'Miktar': float(miktar),
            'Sebep': sebep,
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Iadeler_Hurda", excel_data)

        # İade ise stoğa geri ekle
        if tip == "İade":
            existing = db.fetch_one("SELECT * FROM stok WHERE malzeme = %s", [urun])
            if existing:
                new_miktar = Decimal(str(existing['miktar_kg'])) + miktar
                db.update('stok', {'miktar_kg': new_miktar, 'updated_at': datetime.now()}, 
                         'malzeme = %s', [urun])
            else:
                db.insert('stok', {'malzeme': urun, 'miktar_kg': miktar})

        messagebox.showinfo("Başarılı", "Kayıt eklendi ve Excel'e aktarıldı.")
        combo_iade_urun.set("")
        entry_iade_miktar.delete(0, tk.END)
        entry_iade_sebep.delete(0, tk.END)
        combo_iade_tip.set("")
        guncelle_comboboxlar()
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f5 = ttk.Frame(notebook)
notebook.add(f5, text="İade / Hurda")
tk.Label(f5, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=0, columnspan=2, pady=5)
tk.Label(f5, text="Ürün/Malzeme: ").grid(row=1, column=0, padx=5, pady=5)
combo_iade_urun = ttk.Combobox(f5, values=get_malzeme_ve_urunler(), state="readonly")
combo_iade_urun.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f5, text="Miktar (kg): ").grid(row=2, column=0, padx=5, pady=5)
entry_iade_miktar = tk.Entry(f5)
entry_iade_miktar.grid(row=2, column=1, padx=5, pady=5)
tk.Label(f5, text="Tür: ").grid(row=3, column=0, padx=5, pady=5)
combo_iade_tip = ttk.Combobox(f5, values=["İade", "Hurda"], state="readonly")
combo_iade_tip.grid(row=3, column=1, padx=5, pady=5)
tk.Label(f5, text="Sebep: ").grid(row=4, column=0, padx=5, pady=5)
entry_iade_sebep = tk.Entry(f5)
entry_iade_sebep.grid(row=4, column=1, padx=5, pady=5)
tk.Button(f5, text="Kaydet", command=iade_kaydet).grid(row=5, columnspan=2, pady=10)

# === TAŞ GİDER SEKMESİ ===
tas_gider_turleri = [
    "İŞÇİLİK SGK", "İŞÇİLİK MAAŞ", "İŞ GÜVENLİĞİ", "ÇEVRE DANIŞMANLIK FİRMASI",
    "MADEN MÜHENDİSİ", "SORUMLU YTK", "ORMAN KİRA BEDELİ", "MAPEG KİRA BEDELİ",
    "PATLATMA GİDERİ", "ELEKTRİK", "YEMEK", "MOTORİN", "TAMİR BAKIM GİDERLERİ",
    "YÖNETİM GİDERİ", "VERGİ", "DİĞER"
]

def tas_gider_kaydet():
    try:
        tarih = datetime.strptime(entry_tas_tarih.get(), "%Y-%m-%d").date()
        tip = "Gider"
        aciklama = combo_tas_kategori.get()
        birim = entry_tas_birim.get()
        birim_fiyat = Decimal(str(entry_tas_fiyat.get()))
        miktar = Decimal(str(entry_tas_miktar.get()))
        toplam = birim_fiyat * miktar

        db.insert('tas_gelir_gider', {
            'tarih': tarih,
            'tip': tip,
            'aciklama': aciklama,
            'birim': birim,
            'birim_fiyat': birim_fiyat,
            'miktar': miktar,
            'toplam_tutar': toplam
        })

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Tip': tip,
            'Aciklama': aciklama,
            'Birim': birim,
            'Birim Fiyat': float(birim_fiyat),
            'Miktar': float(miktar),
            'Toplam Tutar': float(toplam),
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Tas_Gelir_Gider", excel_data)

        messagebox.showinfo("Başarılı", "Taş gideri kaydedildi ve Excel'e aktarıldı.")
        entry_tas_tarih.delete(0, tk.END)
        combo_tas_kategori.set("")
        entry_tas_birim.delete(0, tk.END)
        entry_tas_fiyat.delete(0, tk.END)
        entry_tas_miktar.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f6 = ttk.Frame(notebook)
notebook.add(f6, text="Taş Gider")
tk.Label(f6, text="Tarih (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5)
entry_tas_tarih = tk.Entry(f6)
entry_tas_tarih.grid(row=0, column=1, padx=5, pady=5)
tk.Label(f6, text="Gider Türü:").grid(row=1, column=0, padx=5, pady=5)
combo_tas_kategori = ttk.Combobox(f6, values=tas_gider_turleri, state="readonly")
combo_tas_kategori.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f6, text="Birim:").grid(row=2, column=0, padx=5, pady=5)
entry_tas_birim = tk.Entry(f6)
entry_tas_birim.grid(row=2, column=1, padx=5, pady=5)
tk.Label(f6, text="Birim Fiyatı:").grid(row=3, column=0, padx=5, pady=5)
entry_tas_fiyat = tk.Entry(f6)
entry_tas_fiyat.grid(row=3, column=1, padx=5, pady=5)
tk.Label(f6, text="Miktar:").grid(row=4, column=0, padx=5, pady=5)
entry_tas_miktar = tk.Entry(f6)
entry_tas_miktar.grid(row=4, column=1, padx=5, pady=5)
tk.Button(f6, text="Kaydet", command=tas_gider_kaydet).grid(row=5, columnspan=2, pady=10)
tk.Label(f6, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=6, columnspan=2, pady=5)

# === BETON GİDER SEKMESİ ===
beton_gider_turleri = ["ÇİMENTO", "AGREGA", "KATKI"]

def beton_gider_kaydet():
    try:
        tarih = datetime.strptime(entry_beton_tarih.get(), "%Y-%m-%d").date()
        tip = "Gider"
        aciklama = combo_beton_kategori.get()
        birim = entry_beton_birim.get()
        birim_fiyat = Decimal(str(entry_beton_fiyat.get()))
        miktar = Decimal(str(entry_beton_miktar.get()))
        toplam = birim_fiyat * miktar

        db.insert('beton_gelir_gider', {
            'tarih': tarih,
            'tip': tip,
            'aciklama': aciklama,
            'birim': birim,
            'birim_fiyat': birim_fiyat,
            'miktar': miktar,
            'toplam_tutar': toplam
        })

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Tip': tip,
            'Aciklama': aciklama,
            'Birim': birim,
            'Birim Fiyat': float(birim_fiyat),
            'Miktar': float(miktar),
            'Toplam Tutar': float(toplam),
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Beton_Gelir_Gider", excel_data)

        messagebox.showinfo("Başarılı", "Beton gideri kaydedildi ve Excel'e aktarıldı.")
        entry_beton_tarih.delete(0, tk.END)
        combo_beton_kategori.set("")
        entry_beton_birim.delete(0, tk.END)
        entry_beton_fiyat.delete(0, tk.END)
        entry_beton_miktar.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))
        birim_fiyat = float(entry_beton_fiyat.get())
        miktar = float(entry_beton_miktar.get())
        toplam = birim_fiyat * miktar

        db.insert('beton_gelir_gider', {
            'tarih': tarih,
            'tip': tip,
            'aciklama': aciklama,
            'birim': birim,
            'birim_fiyat': birim_fiyat,
            'miktar': miktar,
            'toplam_tutar': toplam
        })

        # Excel kaydı oluştur
        excel_data = {
            'Tarih': tarih.strftime("%Y-%m-%d"),
            'Tip': tip,
            'Aciklama': aciklama,
            'Birim': birim,
            'Birim Fiyat': birim_fiyat,
            'Miktar': miktar,
            'Toplam Tutar': toplam,
            'Kayit Zamani': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        excel_kayit_olustur("Beton_Gelir_Gider", excel_data)

        messagebox.showinfo("Başarılı", "Beton gideri kaydedildi ve Excel'e aktarıldı.")
        entry_beton_tarih.delete(0, tk.END)
        combo_beton_kategori.set("")
        entry_beton_birim.delete(0, tk.END)
        entry_beton_fiyat.delete(0, tk.END)
        entry_beton_miktar.delete(0, tk.END)
    except Exception as e:
        messagebox.showerror("Hata", str(e))

f7 = ttk.Frame(notebook)
notebook.add(f7, text="Beton Gider")
tk.Label(f7, text="Tarih (YYYY-MM-DD):").grid(row=0, column=0, padx=5, pady=5)
entry_beton_tarih = tk.Entry(f7)
entry_beton_tarih.grid(row=0, column=1, padx=5, pady=5)
tk.Label(f7, text="Gider Türü:").grid(row=1, column=0, padx=5, pady=5)
combo_beton_kategori = ttk.Combobox(f7, values=beton_gider_turleri, state="readonly")
combo_beton_kategori.grid(row=1, column=1, padx=5, pady=5)
tk.Label(f7, text="Birim:").grid(row=2, column=0, padx=5, pady=5)
entry_beton_birim = tk.Entry(f7)
entry_beton_birim.grid(row=2, column=1, padx=5, pady=5)
tk.Label(f7, text="Birim Fiyatı:").grid(row=3, column=0, padx=5, pady=5)
entry_beton_fiyat = tk.Entry(f7)
entry_beton_fiyat.grid(row=3, column=1, padx=5, pady=5)
tk.Label(f7, text="Miktar:").grid(row=4, column=0, padx=5, pady=5)
entry_beton_miktar = tk.Entry(f7)
entry_beton_miktar.grid(row=4, column=1, padx=5, pady=5)
tk.Button(f7, text="Kaydet", command=beton_gider_kaydet).grid(row=5, columnspan=2, pady=10)
tk.Label(f7, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").grid(row=6, columnspan=2, pady=5)

# === RAPORLAMA SEKMESİ ===
def raporla():
    try:
        secim = combo_rapor_tipi.get()
        
        if secim == "Günlük":
            # Günlük rapor
            satis_query = """
                SELECT tarih, SUM(net_kar) as toplam_kar
                FROM satislar 
                GROUP BY tarih 
                ORDER BY tarih DESC
                LIMIT 30
            """
            
            tas_query = """
                SELECT tarih, 
                       SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net_tutar
                FROM tas_gelir_gider 
                GROUP BY tarih 
                ORDER BY tarih DESC
                LIMIT 30
            """
            
            beton_query = """
                SELECT tarih, 
                       SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net_tutar
                FROM beton_gelir_gider 
                GROUP BY tarih 
                ORDER BY tarih DESC
                LIMIT 30
            """
        else:
            # Aylık rapor
            satis_query = """
                SELECT DATE_TRUNC('month', tarih) as ay, SUM(net_kar) as toplam_kar
                FROM satislar 
                GROUP BY DATE_TRUNC('month', tarih)
                ORDER BY ay DESC
                LIMIT 12
            """
            
            tas_query = """
                SELECT DATE_TRUNC('month', tarih) as ay, 
                       SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net_tutar
                FROM tas_gelir_gider 
                GROUP BY DATE_TRUNC('month', tarih)
                ORDER BY ay DESC
                LIMIT 12
            """
            
            beton_query = """
                SELECT DATE_TRUNC('month', tarih) as ay, 
                       SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net_tutar
                FROM beton_gelir_gider 
                GROUP BY DATE_TRUNC('month', tarih)
                ORDER BY ay DESC
                LIMIT 12
            """

        satis_data = db.fetch_all(satis_query)
        tas_data = db.fetch_all(tas_query)
        beton_data = db.fetch_all(beton_query)

        # Raporu birleştir ve göster
        liste_rapor.delete(0, tk.END)
        
        # Tüm tarihleri/ayları topla
        all_periods = set()
        for row in satis_data:
            all_periods.add(row[list(row.keys())[0]])
        for row in tas_data:
            all_periods.add(row[list(row.keys())[0]])
        for row in beton_data:
            all_periods.add(row[list(row.keys())[0]])
        
        # Sözlük haline getir
        satis_dict = {row[list(row.keys())[0]]: row[list(row.keys())[1]] or 0 for row in satis_data}
        tas_dict = {row[list(row.keys())[0]]: row[list(row.keys())[1]] or 0 for row in tas_data}
        beton_dict = {row[list(row.keys())[0]]: row[list(row.keys())[1]] or 0 for row in beton_data}
        
        # Sonuçları göster
        for period in sorted(all_periods, reverse=True):
            satis_kar = satis_dict.get(period, 0)
            tas_net = tas_dict.get(period, 0)
            beton_net = beton_dict.get(period, 0)
            toplam_net = satis_kar + tas_net + beton_net
            
            period_str = period.strftime("%Y-%m-%d") if secim == "Günlük" else period.strftime("%Y-%m")
            liste_rapor.insert(tk.END, 
                f"{period_str} ➤ Satış: {satis_kar:.2f} | Taş: {tas_net:.2f} | Beton: {beton_net:.2f} | NET: {toplam_net:.2f} ₺")

    except Exception as e:
        messagebox.showerror("Hata", str(e))

def stok_raporu():
    """Mevcut stok durumunu göster"""
    try:
        stok_data = db.fetch_all("SELECT malzeme, miktar_kg FROM stok WHERE miktar_kg > 0 ORDER BY malzeme")
        
        liste_rapor.delete(0, tk.END)
        liste_rapor.insert(tk.END, "=== MEVCUT STOK DURUMU ===")
        liste_rapor.insert(tk.END, "")
        
        for row in stok_data:
            liste_rapor.insert(tk.END, f"{row['malzeme']}: {row['miktar_kg']:.2f} kg")
            
    except Exception as e:
        messagebox.showerror("Hata", str(e))

def urun_raporu():
    """Tanımlı ürünleri ve reçetelerini göster"""
    try:
        urun_data = db.fetch_all("""
            SELECT urun, malzeme, yuzde 
            FROM urunler 
            ORDER BY urun, malzeme
        """)
        
        liste_rapor.delete(0, tk.END)
        liste_rapor.insert(tk.END, "=== ÜRÜN REÇETELERİ ===")
        liste_rapor.insert(tk.END, "")
        
        current_urun = ""
        for row in urun_data:
            if row['urun'] != current_urun:
                current_urun = row['urun']
                liste_rapor.insert(tk.END, f"📦 {current_urun}:")
            
            liste_rapor.insert(tk.END, f"   • {row['malzeme']}: %{row['yuzde']}")
            
    except Exception as e:
        messagebox.showerror("Hata", str(e))

def excel_dosyalarini_ac():
    """Excel kayıt klasörünü aç"""
    try:
        import os
        import subprocess
        import platform
        
        klasor = "excel_kayitlari"
        if not os.path.exists(klasor):
            messagebox.showwarning("Uyarı", "Excel kayıtları klasörü bulunamadı.")
            return
        
        # İşletim sistemine göre klasörü aç
        if platform.system() == "Windows":
            os.startfile(klasor)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", klasor])
        else:  # Linux
            subprocess.call(["xdg-open", klasor])
            
    except Exception as e:
        messagebox.showerror("Hata", f"Klasör açılamadı: {str(e)}")

f8 = ttk.Frame(notebook)
notebook.add(f8, text="Raporlama")

# Rapor türü seçimi
rapor_frame = tk.Frame(f8)
rapor_frame.pack(pady=10)

tk.Label(rapor_frame, text="Rapor Tipi:").grid(row=0, column=0, padx=5)
combo_rapor_tipi = ttk.Combobox(rapor_frame, values=["Günlük", "Aylık"], state="readonly")
combo_rapor_tipi.set("Günlük")
combo_rapor_tipi.grid(row=0, column=1, padx=5)

# Butonlar
buton_frame = tk.Frame(f8)
buton_frame.pack(pady=5)

tk.Button(buton_frame, text="Gelir-Gider Raporu", command=raporla).grid(row=0, column=0, padx=5)
tk.Button(buton_frame, text="Stok Raporu", command=stok_raporu).grid(row=0, column=1, padx=5)
tk.Button(buton_frame, text="Ürün Raporu", command=urun_raporu).grid(row=0, column=2, padx=5)
tk.Button(buton_frame, text="Excel Kayıtlarını Aç", command=excel_dosyalarini_ac, bg="lightblue").grid(row=0, column=3, padx=5)

# Rapor listesi
liste_rapor = tk.Listbox(f8, width=100, height=20, font=("Consolas", 9))
liste_rapor.pack(padx=10, pady=10, fill="both", expand=True)

# Scrollbar ekle
scrollbar = tk.Scrollbar(f8)
scrollbar.pack(side="right", fill="y")
liste_rapor.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=liste_rapor.yview)

tk.Label(f8, text="TURKCE KARAKTER KULLANMAYIN!", fg="red").pack(pady=5)

# === GENEL EXCEL RAPORU FONKSİYONU ===
def excel_raporu_olustur():
    """Tüm verileri Excel dosyasına kaydet"""
    try:
        # Dosya adı oluştur
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"beton_takip_genel_raporu_{timestamp}.xlsx"
        
        # Excel workbook oluştur
        wb = openpyxl.Workbook()
        
        # Varsayılan sheet'i sil
        wb.remove(wb.active)
        
        # Stil tanımlamaları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def format_sheet(ws, df):
            """Sheet'i formatla"""
            # Header'ları formatla
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Tüm hücrelere border ekle
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
            
            # Sütun genişliklerini ayarla
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 1. STOK RAPORU
        stok_data = db.fetch_all("SELECT malzeme, miktar_kg, created_at, updated_at FROM stok ORDER BY malzeme")
        if stok_data:
            df_stok = pd.DataFrame(stok_data)
            ws_stok = wb.create_sheet("Stok Durumu")
            
            for row in dataframe_to_rows(df_stok, index=False, header=True):
                ws_stok.append(row)
            
            format_sheet(ws_stok, df_stok)
        
        # 2. ALIŞLAR RAPORU
        alis_data = db.fetch_all("""
            SELECT malzeme, miktar_kg, birim_fiyat, toplam_tutar, tarih, created_at 
            FROM alislar 
            ORDER BY tarih DESC, created_at DESC
        """)
        if alis_data:
            df_alis = pd.DataFrame(alis_data)
            ws_alis = wb.create_sheet("Alışlar")
            
            for row in dataframe_to_rows(df_alis, index=False, header=True):
                ws_alis.append(row)
            
            format_sheet(ws_alis, df_alis)
            
            # Toplam satırı ekle
            total_row = ws_alis.max_row + 2
            ws_alis[f'C{total_row}'] = "TOPLAM:"
            ws_alis[f'D{total_row}'] = f"=SUM(D2:D{ws_alis.max_row-1})"
            ws_alis[f'C{total_row}'].font = Font(bold=True)
            ws_alis[f'D{total_row}'].font = Font(bold=True)
        
        # 3. ÜRÜNLER VE REÇETELER
        urun_data = db.fetch_all("SELECT urun, malzeme, yuzde, created_at FROM urunler ORDER BY urun, malzeme")
        if urun_data:
            df_urun = pd.DataFrame(urun_data)
            ws_urun = wb.create_sheet("Ürün Reçeteleri")
            
            for row in dataframe_to_rows(df_urun, index=False, header=True):
                ws_urun.append(row)
            
            format_sheet(ws_urun, df_urun)
        
        # 4. ÜRETİMLER RAPORU
        uretim_data = db.fetch_all("SELECT urun, gramaj_kg, tarih, created_at FROM uretimler ORDER BY tarih DESC")
        if uretim_data:
            df_uretim = pd.DataFrame(uretim_data)
            ws_uretim = wb.create_sheet("Üretimler")
            
            for row in dataframe_to_rows(df_uretim, index=False, header=True):
                ws_uretim.append(row)
            
            format_sheet(ws_uretim, df_uretim)
            
            # Toplam üretim
            total_row = ws_uretim.max_row + 2
            ws_uretim[f'A{total_row}'] = "TOPLAM ÜRETİM:"
            ws_uretim[f'B{total_row}'] = f"=SUM(B2:B{ws_uretim.max_row-1})"
            ws_uretim[f'A{total_row}'].font = Font(bold=True)
            ws_uretim[f'B{total_row}'].font = Font(bold=True)
        
        # 5. SATIŞLAR RAPORU
        satis_data = db.fetch_all("""
            SELECT urun, musteri, miktar_kg, satis_fiyat, toplam_satis, net_kar, tarih, created_at 
            FROM satislar 
            ORDER BY tarih DESC, created_at DESC
        """)
        if satis_data:
            df_satis = pd.DataFrame(satis_data)
            ws_satis = wb.create_sheet("Satışlar")
            
            for row in dataframe_to_rows(df_satis, index=False, header=True):
                ws_satis.append(row)
            
            format_sheet(ws_satis, df_satis)
            
            # Toplam satırları
            total_row = ws_satis.max_row + 2
            ws_satis[f'D{total_row}'] = "TOPLAM:"
            ws_satis[f'E{total_row}'] = f"=SUM(E2:E{ws_satis.max_row-1})"  # Toplam satış
            ws_satis[f'F{total_row}'] = f"=SUM(F2:F{ws_satis.max_row-1})"  # Toplam kar
            ws_satis[f'D{total_row}'].font = Font(bold=True)
            ws_satis[f'E{total_row}'].font = Font(bold=True)
            ws_satis[f'F{total_row}'].font = Font(bold=True)
        
        # 6. İADELER/HURDA RAPORU
        iade_data = db.fetch_all("SELECT tarih, tip, urun, miktar, sebep, created_at FROM iadeler ORDER BY tarih DESC")
        if iade_data:
            df_iade = pd.DataFrame(iade_data)
            ws_iade = wb.create_sheet("İadeler-Hurda")
            
            for row in dataframe_to_rows(df_iade, index=False, header=True):
                ws_iade.append(row)
            
            format_sheet(ws_iade, df_iade)
        
        # 7. TAŞ GELİR-GİDER RAPORU
        tas_data = db.fetch_all("""
            SELECT tarih, tip, aciklama, birim, birim_fiyat, miktar, toplam_tutar, created_at 
            FROM tas_gelir_gider 
            ORDER BY tarih DESC, created_at DESC
        """)
        if tas_data:
            df_tas = pd.DataFrame(tas_data)
            ws_tas = wb.create_sheet("Taş Gelir-Gider")
            
            for row in dataframe_to_rows(df_tas, index=False, header=True):
                ws_tas.append(row)
            
            format_sheet(ws_tas, df_tas)
        
        # 8. BETON GELİR-GİDER RAPORU
        beton_data = db.fetch_all("""
            SELECT tarih, tip, aciklama, birim, birim_fiyat, miktar, toplam_tutar, created_at 
            FROM beton_gelir_gider 
            ORDER BY tarih DESC, created_at DESC
        """)
        if beton_data:
            df_beton = pd.DataFrame(beton_data)
            ws_beton = wb.create_sheet("Beton Gelir-Gider")
            
            for row in dataframe_to_rows(df_beton, index=False, header=True):
                ws_beton.append(row)
            
            format_sheet(ws_beton, df_beton)
        
        # 9. ÖZET RAPORU
        ws_ozet = wb.create_sheet("Özet Rapor")
        wb.active = ws_ozet  # Özet raporu aktif sheet yap
        
        # Özet verilerini hesapla
        ozet_data = []
        
        # Toplam stok
        toplam_stok = db.fetch_one("SELECT SUM(miktar_kg) as toplam FROM stok")
        ozet_data.append(["Toplam Stok (kg)", toplam_stok['toplam'] if toplam_stok['toplam'] else 0])
        
        # Toplam alış tutarı
        toplam_alis = db.fetch_one("SELECT SUM(toplam_tutar) as toplam FROM alislar")
        ozet_data.append(["Toplam Alış Tutarı (TL)", toplam_alis['toplam'] if toplam_alis['toplam'] else 0])
        
        # Toplam üretim
        toplam_uretim = db.fetch_one("SELECT SUM(gramaj_kg) as toplam FROM uretimler")
        ozet_data.append(["Toplam Üretim (kg)", toplam_uretim['toplam'] if toplam_uretim['toplam'] else 0])
        
        # Toplam satış tutarı
        toplam_satis_tutar = db.fetch_one("SELECT SUM(toplam_satis) as toplam FROM satislar")
        ozet_data.append(["Toplam Satış Tutarı (TL)", toplam_satis_tutar['toplam'] if toplam_satis_tutar['toplam'] else 0])
        
        # Toplam net kar
        toplam_kar = db.fetch_one("SELECT SUM(net_kar) as toplam FROM satislar")
        ozet_data.append(["Toplam Net Kar (TL)", toplam_kar['toplam'] if toplam_kar['toplam'] else 0])
        
        # Taş gelir-gider net
        tas_net = db.fetch_one("""
            SELECT SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net 
            FROM tas_gelir_gider
        """)
        ozet_data.append(["Taş İşleri Net (TL)", tas_net['net'] if tas_net['net'] else 0])
        
        # Beton gelir-gider net
        beton_net = db.fetch_one("""
            SELECT SUM(CASE WHEN tip = 'Gelir' THEN toplam_tutar ELSE -toplam_tutar END) as net 
            FROM beton_gelir_gider
        """)
        ozet_data.append(["Beton İşleri Net (TL)", beton_net['net'] if beton_net['net'] else 0])
        
        # Özet tablosunu oluştur
        headers = ["Kategori", "Değer"]
        ws_ozet.append(headers)
        
        for row in ozet_data:
            ws_ozet.append(row)
        
        # Özet raporu formatla
        for cell in ws_ozet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row in ws_ozet.iter_rows(min_row=2):
            row[0].font = Font(bold=True)
            for cell in row:
                cell.border = border
        
        # Sütun genişliklerini ayarla
        ws_ozet.column_dimensions['A'].width = 25
        ws_ozet.column_dimensions['B'].width = 20
        
        # Rapor oluşturma tarihi ekle
        ws_ozet[f'A{len(ozet_data) + 3}'] = "Rapor Tarihi:"
        ws_ozet[f'B{len(ozet_data) + 3}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_ozet[f'A{len(ozet_data) + 3}'].font = Font(bold=True)
        
        # Excel dosyasını kaydet
        wb.save(filename)
        
        messagebox.showinfo("Başarılı", f"Genel Excel raporu oluşturuldu: {filename}")
        
        # Dosyayı açmak isteyip istemediğini sor
        result = messagebox.askyesno("Dosyayı Aç", "Excel dosyasını şimdi açmak istiyor musunuz?")
        if result:
            os.startfile(filename)  # Windows için
            
    except Exception as e:
        messagebox.showerror("Hata", f"Excel raporu oluşturulurken hata: {str(e)}")

# Genel Excel raporu butonu
tk.Button(buton_frame, text="Genel Excel Raporu", command=excel_raporu_olustur, bg="lightgreen").grid(row=1, column=0, columnspan=4, pady=5)

# === VERİTABANI YÖNETIM SEKMESİ ===
def veritabani_yedekle():
    """Veritabanını yedekle (CSV formatında)"""
    try:
        import os
        from datetime import datetime
        
        # Yedek klasörü oluştur
        backup_dir = "veritabani_yedekleri"
        os.makedirs(backup_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Tüm tabloları yedekle
        tables = ['stok', 'alislar', 'urunler', 'uretimler', 'satislar', 
                 'iadeler', 'tas_gelir_gider', 'beton_gelir_gider']
        
        for table in tables:
            data = db.fetch_all(f"SELECT * FROM {table}")
            if data:
                df = pd.DataFrame(data)
                df.to_csv(f"{backup_dir}/{table}_{timestamp}.csv", index=False, encoding='utf-8')
        
        messagebox.showinfo("Başarılı", f"Veritabanı {backup_dir} klasörüne yedeklendi.")
        
    except Exception as e:
        messagebox.showerror("Hata", f"Yedekleme hatası: {str(e)}")

def veritabani_temizle():
    """Tüm tabloları temizle (dikkatli kullanın!)"""
    result = messagebox.askyesno("Uyarı", 
        "TÜM VERİLER SİLİNECEK!\n\nBu işlem geri alınamaz. Devam etmek istediğinizden emin misiniz?")
    
    if result:
        result2 = messagebox.askyesno("Son Uyarı", 
            "SON UYARI: Tüm veriler kalıcı olarak silinecek!\n\nGerçekten devam etmek istiyor musunuz?")
        
        if result2:
            try:
                tables = ['satislar', 'uretimler', 'iadeler', 'tas_gelir_gider', 
                         'beton_gelir_gider', 'urunler', 'alislar', 'stok']
                
                for table in tables:
                    db.execute_query(f"DELETE FROM {table}")
                
                messagebox.showinfo("Tamamlandı", "Tüm veriler silindi.")
                guncelle_comboboxlar()
                
            except Exception as e:
                messagebox.showerror("Hata", f"Temizleme hatası: {str(e)}")

f9 = ttk.Frame(notebook)
notebook.add(f9, text="Veritabanı Yönetimi")

# Bağlantı bilgileri
info_frame = tk.LabelFrame(f9, text="Bağlantı Bilgileri", padx=10, pady=10)
info_frame.pack(padx=10, pady=10, fill="x")

tk.Label(info_frame, text=f"Sunucu: {db.config.host}:{db.config.port}").pack(anchor="w")
tk.Label(info_frame, text=f"Veritabanı: {db.config.database}").pack(anchor="w")
tk.Label(info_frame, text=f"Kullanıcı: {db.config.username}").pack(anchor="w")

# Yönetim butonları
yonetim_frame = tk.LabelFrame(f9, text="Veritabanı İşlemleri", padx=10, pady=10)
yonetim_frame.pack(padx=10, pady=10, fill="x")

tk.Button(yonetim_frame, text="Veritabanını Yedekle", command=veritabani_yedekle, 
         bg="lightgreen").pack(pady=5, fill="x")

tk.Button(yonetim_frame, text="Tüm Verileri Temizle", command=veritabani_temizle, 
         bg="lightcoral", fg="white").pack(pady=5, fill="x")

# Excel kayıtları yönetimi
excel_frame = tk.LabelFrame(f9, text="Excel Kayıtları", padx=10, pady=10)
excel_frame.pack(padx=10, pady=10, fill="x")

tk.Label(excel_frame, text="Her işlem otomatik olarak ayrı Excel dosyalarına kaydediliyor:").pack(anchor="w")
tk.Label(excel_frame, text="• Alışlar: excel_kayitlari/Alislar.xlsx").pack(anchor="w", padx=20)
tk.Label(excel_frame, text="• Satışlar: excel_kayitlari/Satislar.xlsx").pack(anchor="w", padx=20)
tk.Label(excel_frame, text="• Üretimler: excel_kayitlari/Uretimler.xlsx").pack(anchor="w", padx=20)
tk.Label(excel_frame, text="• Giderler: excel_kayitlari/Tas_Gelir_Gider.xlsx ve Beton_Gelir_Gider.xlsx").pack(anchor="w", padx=20)
tk.Label(excel_frame, text="• İade/Hurda: excel_kayitlari/Iadeler_Hurda.xlsx").pack(anchor="w", padx=20)
tk.Label(excel_frame, text="• Ürün Reçeteleri: excel_kayitlari/Urun_Receteleri.xlsx").pack(anchor="w", padx=20)

tk.Button(excel_frame, text="Excel Kayıtları Klasörünü Aç", command=excel_dosyalarini_ac, 
         bg="lightblue").pack(pady=10, fill="x")

# Uyarı
tk.Label(f9, text="⚠️ Veritabanı işlemlerini dikkatli kullanın!", 
         fg="red", font=("Arial", 10, "bold")).pack(pady=10)

# === YARDIMCI FONKSİYONLAR ===
def guncelle_comboboxlar():
    """Tüm combobox'ları güncelle"""
    try:
        # Malzeme listelerini güncelle
        malzemeler = get_malzemeler()
        combo_urun_malzeme['values'] = malzemeler
        
        # Ürün listelerini güncelle
        urunler = get_urunler()
        combo_uretim_urun['values'] = urunler
        combo_satis_urun['values'] = urunler
        
        # Malzeme + ürün listesini güncelle
        malzeme_ve_urunler = get_malzeme_ve_urunler()
        combo_iade_urun['values'] = malzeme_ve_urunler
        
    except Exception as e:
        print(f"Combobox güncelleme hatası: {e}")

# Program başlatıldığında combobox'ları güncelle
guncelle_comboboxlar()

# === PROGRAM BAŞLAT ===
if __name__ == "__main__":
    try:
        # Excel kayıtları klasörünü oluştur
        os.makedirs("excel_kayitlari", exist_ok=True)
        print("Excel kayıtları klasörü hazır: excel_kayitlari/")
        
        root.mainloop()
    finally:
        # Veritabanı bağlantısını kapat
        if db and db.connection:
            db.connection.close()